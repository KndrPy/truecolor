#!/usr/bin/env python3
"""Inspect the private ISSA workbook without exporting subject-level records."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--preview-rows", type=int, default=5)
    parser.add_argument("--header-scan-rows", type=int, default=15)
    return parser.parse_args()


def json_safe(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def main() -> None:
    args = parse_args()
    if not args.input.exists():
        raise FileNotFoundError(args.input)

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    report: dict[str, object] = {
        "input": str(args.input),
        "size_bytes": args.input.stat().st_size,
        "sheets": [],
    }

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        scan = []
        for row_index, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(args.header_scan_rows, worksheet.max_row),
                values_only=True,
            ),
            start=1,
        ):
            values = [json_safe(value) for value in row]
            nonempty = sum(value not in (None, "") for value in values)
            scan.append({
                "row": row_index,
                "nonempty_cells": nonempty,
                "values": values,
            })

        sheet_report: dict[str, object] = {
            "sheet": sheet_name,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "header_scan": scan,
        }

        try:
            preview = pd.read_excel(
                args.input,
                sheet_name=sheet_name,
                nrows=args.preview_rows,
            )
            sheet_report["default_header_columns"] = [
                str(column) for column in preview.columns
            ]
            sheet_report["default_header_preview"] = [
                {str(key): json_safe(value) for key, value in row.items()}
                for row in preview.to_dict(orient="records")
            ]
        except Exception as exc:
            sheet_report["preview_error"] = f"{type(exc).__name__}: {exc}"

        report["sheets"].append(sheet_report)

    rendered = json.dumps(report, indent=2, ensure_ascii=False)
    print(rendered)

    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(rendered + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
