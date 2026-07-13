from __future__ import annotations

import csv
import datetime as dt
import hashlib
import math

from io import BytesIO

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import (
    get_column_letter,
)


class TabularParseError(RuntimeError):
    def __init__(
        self,
        code: str,
        message: str,
    ) -> None:
        super().__init__(message)
        self.code = code


def canonical_hash(
    *values: object,
) -> str:
    payload = "\n".join(
        str(value)
        for value in values
    ).encode("utf-8")

    return hashlib.sha256(
        payload
    ).hexdigest()


def value_type(
    value: Any,
    *,
    formula: bool = False,
) -> str:
    if formula:
        return "FORMULA"

    if value is None or value == "":
        return "EMPTY"

    if isinstance(value, bool):
        return "BOOLEAN"

    if isinstance(value, int):
        return "INTEGER"

    if isinstance(value, float):
        return "NUMBER"

    if isinstance(value, dt.datetime):
        return "DATETIME"

    if isinstance(value, dt.date):
        return "DATE"

    return "STRING"


def canonical_value(
    value: Any,
) -> Any:
    if isinstance(value, dt.datetime):
        return value.isoformat()

    if isinstance(value, dt.date):
        return value.isoformat()

    if (
        isinstance(value, float)
        and not math.isfinite(value)
    ):
        raise TabularParseError(
            "NONFINITE_NUMBER",
            repr(value),
        )

    return value


def source_record(
    *,
    source_start: int | None,
    source_end: int | None,
    table_index: int,
    sheet_name: str | None,
    row_index: int,
    column_index: int,
    cell_reference: str,
) -> dict[str, Any]:
    return {
        "page_number": None,
        "source_start": (
            source_start
        ),
        "source_end": source_end,
        "bbox": None,
        "xml_path": None,
        "document_part": None,
        "paragraph_index": None,
        "table_index": table_index,
        "sheet_name": sheet_name,
        "row_index": row_index,
        "column_index": (
            column_index
        ),
        "cell_reference": (
            cell_reference
        ),
    }


def build_table_document(
    *,
    artifact_id: str,
    source_kind: str,
    name: str | None,
    table_index: int,
    rows: list[list[Any]],
    source_offsets: list[
        list[tuple[int, int] | None]
    ] | None,
) -> tuple[
    str,
    list[dict[str, Any]],
    dict[str, Any],
]:
    pieces: list[str] = []
    segments = []
    table_rows = []
    maximum_columns = 0

    for row_index, row in enumerate(
        rows,
        start=1,
    ):
        normalized_cells = []
        row_text_values = []

        for column_index, raw_value in enumerate(
            row,
            start=1,
        ):
            formula = (
                isinstance(
                    raw_value,
                    str,
                )
                and raw_value.startswith("=")
            )

            normalized = canonical_value(
                raw_value
            )

            text_value = (
                ""
                if normalized is None
                else str(normalized)
            )

            cell_reference = (
                f"{get_column_letter(column_index)}"
                f"{row_index}"
            )

            normalized_cells.append(
                {
                    "column_index": (
                        column_index
                    ),
                    "cell_reference": (
                        cell_reference
                    ),
                    "value": normalized,
                    "value_type": value_type(
                        raw_value,
                        formula=formula,
                    ),
                }
            )

            row_text_values.append(
                text_value
            )

            if text_value:
                if pieces:
                    pieces.append("\n")

                start = sum(
                    len(piece)
                    for piece in pieces
                )

                pieces.append(
                    text_value
                )

                end = (
                    start
                    + len(text_value)
                )

                offset = None

                if (
                    source_offsets
                    is not None
                    and row_index - 1
                    < len(source_offsets)
                    and column_index - 1
                    < len(
                        source_offsets[
                            row_index - 1
                        ]
                    )
                ):
                    offset = (
                        source_offsets[
                            row_index - 1
                        ][
                            column_index - 1
                        ]
                    )

                source_start = (
                    offset[0]
                    if offset is not None
                    else None
                )

                source_end = (
                    offset[1]
                    if offset is not None
                    else None
                )

                source = source_record(
                    source_start=(
                        source_start
                    ),
                    source_end=(
                        source_end
                    ),
                    table_index=(
                        table_index
                    ),
                    sheet_name=name,
                    row_index=row_index,
                    column_index=(
                        column_index
                    ),
                    cell_reference=(
                        cell_reference
                    ),
                )

                segment_hash = (
                    canonical_hash(
                        artifact_id,
                        source_kind,
                        name,
                        table_index,
                        row_index,
                        column_index,
                        text_value,
                    )
                )

                segments.append(
                    {
                        "segment_id": (
                            f"segment:"
                            f"{segment_hash}"
                        ),
                        "kind": (
                            "TABLE_CELL"
                        ),
                        "text": text_value,
                        "canonical_start": (
                            start
                        ),
                        "canonical_end": end,
                        "source": source,
                    }
                )

        maximum_columns = max(
            maximum_columns,
            len(normalized_cells),
        )

        table_rows.append(
            {
                "row_index": row_index,
                "cells": (
                    normalized_cells
                ),
            }
        )

    table_hash = canonical_hash(
        artifact_id,
        source_kind,
        name,
        table_index,
        table_rows,
    )

    table = {
        "table_id": (
            f"table:{table_hash}"
        ),
        "source_kind": source_kind,
        "name": name,
        "table_index": table_index,
        "row_count": len(
            table_rows
        ),
        "column_count": (
            maximum_columns
        ),
        "rows": table_rows,
    }

    return (
        "".join(pieces),
        segments,
        table,
    )


def parse_delimited(
    path: Path,
    *,
    artifact_id: str,
    delimiter: str,
    source_kind: str,
) -> tuple[
    str,
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    try:
        source_text = (
            path.read_bytes()
            .decode("utf-8-sig")
        )
    except UnicodeDecodeError as exc:
        raise TabularParseError(
            "DELIMITED_SOURCE_NOT_UTF8",
            str(path),
        ) from exc

    rows = list(
        csv.reader(
            source_text.splitlines(),
            delimiter=delimiter,
        )
    )

    offsets: list[
        list[tuple[int, int] | None]
    ] = []

    cursor = 0

    for row in rows:
        row_offsets = []

        for value in row:
            position = source_text.find(
                value,
                cursor,
            )

            if position < 0:
                row_offsets.append(None)
            else:
                end = position + len(value)

                row_offsets.append(
                    (
                        position,
                        end,
                    )
                )

                cursor = end

        offsets.append(
            row_offsets
        )

    text, segments, table = (
        build_table_document(
            artifact_id=artifact_id,
            source_kind=source_kind,
            name=path.name,
            table_index=0,
            rows=rows,
            source_offsets=offsets,
        )
    )

    return (
        text,
        segments,
        [table],
    )


def parse_xlsx(
    path: Path,
    *,
    artifact_id: str,
) -> tuple[
    str,
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    try:
        workbook = load_workbook(
            filename=BytesIO(
                path.read_bytes()
            ),
            read_only=True,
            data_only=False,
        )
    except Exception as exc:
        raise TabularParseError(
            "XLSX_OPEN_FAILED",
            str(exc),
        ) from exc

    document_pieces = []
    all_segments = []
    tables = []

    try:
        for table_index, sheet in enumerate(
            workbook.worksheets
        ):
            rows = [
                [
                    cell.value
                    for cell in row
                ]
                for row in sheet.iter_rows()
            ]

            text, segments, table = (
                build_table_document(
                    artifact_id=(
                        artifact_id
                    ),
                    source_kind="XLSX",
                    name=sheet.title,
                    table_index=(
                        table_index
                    ),
                    rows=rows,
                    source_offsets=None,
                )
            )

            if text:
                prefix = ""

                if document_pieces:
                    prefix = "\n\f\n"
                    document_pieces.append(
                        prefix
                    )

                base_offset = sum(
                    len(piece)
                    for piece
                    in document_pieces
                )

                for segment in segments:
                    segment[
                        "canonical_start"
                    ] += base_offset

                    segment[
                        "canonical_end"
                    ] += base_offset

                document_pieces.append(
                    text
                )

            all_segments.extend(
                segments
            )

            tables.append(
                table
            )

    finally:
        workbook.close()

    return (
        "".join(document_pieces),
        all_segments,
        tables,
    )
